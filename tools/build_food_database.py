"""日本食品標準成分表(八訂)増補2023年をアプリ用JSONとMarkdownに変換する。

出典: 文部科学省 日本食品標準成分表(八訂)増補2023年 第2章(データ)
https://www.mext.go.jp/a_menu/syokuhinseibun/mext_00001.html
"""

import json
import re
import sys
import urllib.request
from datetime import date
from pathlib import Path

import openpyxl

NUMBER_PATTERN = re.compile(r"-?\d+(?:\.\d+)?")
SHEET_PREFIX_PATTERN = re.compile(r"^\d+")

SOURCE_URL = (
    "https://www.mext.go.jp/content/20260327-mxt_kagsei-mext-000029402_02.xlsx"
)
SOURCE_LABEL = "日本食品標準成分表(八訂)増補2023年(文部科学省)"

REPO_ROOT = Path(__file__).resolve().parent.parent
CACHE_PATH = REPO_ROOT / "tools" / "cache" / "mext_seibun.xlsx"
JSON_PATH = REPO_ROOT / "data" / "mext-foods.json"
MARKDOWN_DIR = REPO_ROOT / "docs" / "成分表"

MAIN_SHEET = "表全体"
IDENTIFIER_ROW = 12  # 成分識別子(ENERC_KCAL 等)が並ぶ行(1始まり)
FIRST_DATA_ROW = 13
COL_GROUP_CODE = 0
COL_FOOD_CODE = 1
COL_FOOD_NAME = 3

# 使用する成分識別子。列位置は固定せずこの識別子から解決する。
IDENTIFIERS = {
    "kcal": "ENERC_KCAL",
    "protein": "PROT-",
    "fat": "FAT-",
    "fiber": "FIB-",
    "carb": "CHOCDF-",
    "salt": "NACL_EQ",
}


def parse_value(raw):
    """成分表のセル値を数値化する。

    成分表の数値欄には測定状況を示す記号が混在する:
      Tr / (Tr) 微量、- 未測定、(9.7) 推計値、14.0† 注記付き。
    いずれも計算を止めずに進めるため 0 か数値へ落とす。
    """
    if raw is None:
        return 0.0
    if isinstance(raw, (int, float)):
        return float(raw)

    text = str(raw).strip()
    if text == "":
        return 0.0
    if "Tr" in text:
        return 0.0

    match = NUMBER_PATTERN.search(text)
    if match is None:
        return 0.0
    return float(match.group())


def _is_not_measured(raw):
    if raw is None:
        return True
    if isinstance(raw, (int, float)):
        return False
    return str(raw).strip() in ("", "-")


def calc_sugar(carb_raw, fiber_raw):
    """糖質を「炭水化物 − 食物繊維総量」で算出する。

    食物繊維が未測定の品目では減算せず炭水化物をそのまま採用する。
    食物繊維が炭水化物を上回る品目(こんにゃく等)は 0 にクランプする。
    """
    carb = parse_value(carb_raw)
    if _is_not_measured(fiber_raw):
        return round(carb, 1)
    sugar = carb - parse_value(fiber_raw)
    return round(max(sugar, 0.0), 1)


def normalize_name(raw):
    """食品名の全角スペースを半角に直し、連続する空白を1つにまとめる。"""
    text = str(raw).replace("　", " ")
    return re.sub(r"\s+", " ", text).strip()


def group_name_from_sheet(sheet_name):
    """シート名(例 '1穀類')の先頭の食品群番号を除いて群名を返す。"""
    return SHEET_PREFIX_PATTERN.sub("", sheet_name).strip()


def download_source_if_needed():
    """入力xlsxがローカルに無ければ文科省サイトから取得する。"""
    if CACHE_PATH.exists():
        return CACHE_PATH
    CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    print(f"成分表をダウンロードしています: {SOURCE_URL}")
    urllib.request.urlretrieve(SOURCE_URL, CACHE_PATH)
    return CACHE_PATH


def resolve_columns(identifier_row):
    """成分識別子の行から、必要な成分の列インデックスを解決する。

    文科省が列を追加・並び替えても壊れないよう、列位置は固定値にしない。
    """
    cells = [str(c).strip() if c is not None else "" for c in identifier_row]
    columns = {}
    for key, identifier in IDENTIFIERS.items():
        if identifier not in cells:
            raise SystemExit(
                f"成分識別子 '{identifier}' が{IDENTIFIER_ROW}行目に見つかりません。"
                "成分表の書式が変更された可能性があります。"
            )
        columns[key] = cells.index(identifier)
    return columns


def load_group_names(workbook):
    """シート名から食品群コードと群名の対応表を作る('1穀類' -> {'01': '穀類'})。"""
    groups = {}
    for sheet_name in workbook.sheetnames:
        if sheet_name == MAIN_SHEET:
            continue
        match = SHEET_PREFIX_PATTERN.match(sheet_name)
        if match is None:
            continue
        code = match.group().zfill(2)
        groups[code] = group_name_from_sheet(sheet_name)
    return groups


def read_foods(xlsx_path):
    """xlsxを読み、食品レコードのリストを返す。"""
    workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    group_names = load_group_names(workbook)
    sheet = workbook[MAIN_SHEET]
    rows = list(sheet.iter_rows(values_only=True))
    columns = resolve_columns(rows[IDENTIFIER_ROW - 1])

    foods = []
    for row in rows[FIRST_DATA_ROW - 1:]:
        code = row[COL_FOOD_CODE]
        if code in (None, ""):
            continue
        group_code = str(row[COL_GROUP_CODE]).zfill(2)
        carb_raw = row[columns["carb"]]
        fiber_raw = row[columns["fiber"]]
        foods.append(
            {
                "code": str(code),
                "group": group_names.get(group_code, group_code),
                "name": normalize_name(row[COL_FOOD_NAME]),
                "per100g": {
                    "kcal": round(parse_value(row[columns["kcal"]]), 1),
                    "protein": round(parse_value(row[columns["protein"]]), 1),
                    "fat": round(parse_value(row[columns["fat"]]), 1),
                    "carb": calc_sugar(carb_raw, fiber_raw),
                    "salt": round(parse_value(row[columns["salt"]]), 1),
                },
                # Markdown出力でのみ使う。JSONには含めない。
                "_fiber": round(parse_value(fiber_raw), 1),
                "_carbohydrate": round(parse_value(carb_raw), 1),
                "_group_code": group_code,
            }
        )
    return foods


def write_json(foods):
    """アプリが読むJSONを書き出す(内部用の _ 始まりのキーは除く)。"""
    payload = [
        {
            "code": food["code"],
            "group": food["group"],
            "name": food["name"],
            "per100g": food["per100g"],
        }
        for food in foods
    ]
    JSON_PATH.parent.mkdir(parents=True, exist_ok=True)
    with JSON_PATH.open("w", encoding="utf-8", newline="\n") as f:
        json.dump(payload, f, ensure_ascii=False, separators=(",", ":"))
        f.write("\n")
    print(f"書き出しました: {JSON_PATH} ({len(payload)}品目)")


def write_markdown(foods):
    """食品群ごとにObsidianで読めるMarkdownの表を書き出す。"""
    MARKDOWN_DIR.mkdir(parents=True, exist_ok=True)
    by_group = {}
    for food in foods:
        by_group.setdefault(food["_group_code"], []).append(food)

    generated_on = date.today().isoformat()
    for group_code in sorted(by_group):
        items = by_group[group_code]
        group_name = items[0]["group"]
        lines = [
            f"# {int(group_code)} {group_name}",
            "",
            f"出典: {SOURCE_LABEL}",
            f"生成日: {generated_on} / 品目数: {len(items)}",
            "",
            "すべて可食部100gあたりの値。糖質は「炭水化物 − 食物繊維総量」で算出。",
            "",
            "| 食品番号 | 食品名 | kcal | タンパク質 | 脂質 | 糖質 | 食物繊維 | 炭水化物 | 塩分 |",
            "|---|---|---|---|---|---|---|---|---|",
        ]
        for food in items:
            n = food["per100g"]
            lines.append(
                f"| {food['code']} | {food['name']} | {n['kcal']:g} | {n['protein']:g} | "
                f"{n['fat']:g} | {n['carb']:g} | {food['_fiber']:g} | "
                f"{food['_carbohydrate']:g} | {n['salt']:g} |"
            )
        lines.append("")

        path = MARKDOWN_DIR / f"{group_code}_{group_name}.md"
        path.write_text("\n".join(lines), encoding="utf-8", newline="\n")
    print(f"書き出しました: {MARKDOWN_DIR} ({len(by_group)}ファイル)")


def verify(foods):
    """生成結果を検証し、異常があれば標準エラーに警告を出す。"""
    warnings = []
    total = len(foods)
    if total < 2000:
        warnings.append(f"品目数が想定より少ないです: {total}件(期待: 2000件以上)")

    zero_kcal = sum(1 for f in foods if f["per100g"]["kcal"] == 0)
    if total and zero_kcal / total > 0.10:
        warnings.append(
            f"kcalが0の品目が多すぎます: {zero_kcal}件({zero_kcal / total:.1%})"
        )

    too_large = [f for f in foods if f["per100g"]["kcal"] > 1000]
    if too_large:
        warnings.append(
            f"kcalが1000を超える品目があります: {len(too_large)}件"
            f"(例: {too_large[0]['name']} {too_large[0]['per100g']['kcal']})"
        )

    clamped = sum(
        1 for f in foods if f["per100g"]["carb"] == 0 and f["_carbohydrate"] < f["_fiber"]
    )
    if clamped > 100:
        warnings.append(f"糖質が負になりクランプされた品目が多すぎます: {clamped}件")

    print(f"検証: {total}品目 / kcal=0が{zero_kcal}件 / 糖質クランプが{clamped}件")
    for warning in warnings:
        print(f"警告: {warning}", file=sys.stderr)
    return not warnings


def main():
    xlsx_path = download_source_if_needed()
    foods = read_foods(xlsx_path)
    write_json(foods)
    write_markdown(foods)
    verify(foods)


if __name__ == "__main__":
    main()
